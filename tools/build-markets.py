# -*- coding: utf-8 -*-
"""
Genera data/markets.js a partir de 'Updated City_County Unit List.xlsx'.

DETECCIÓN DE NIVELES — se explica en detalle porque no es obvia:
El Excel no tiene indentación, columnas de nivel ni nada que marque la
jerarquía. Adivinar por el texto falla:
  · "New Market" (AL) y "Haymarket" (VA) son CIUDADES, no mercados.
  · "Washington" es una CIUDAD en Pennsylvania y en DC, no un estado.
La señal fiable es estructural:
  · ZIP        -> formato #####  o  #####-####
  · CONDADO    -> termina en " County" / " Parish" / " city"
  · lo demás   -> MERCADO si le sigue un condado, CIUDAD si le siguen ZIPs
Los títulos de sección ("Ohio Properties", "Properties in the KC Market")
caen fuera de las dos ramas y se descartan solos, porque no les sigue ni un
condado ni un ZIP.

MERCADOS QUE CRUZAN ESTADOS
Kansas City (KS+MO), Chattanooga (GA+TN), Columbia-St Louis (IL+MO) y Ohio
(OH+KY) son UN solo mercado del negocio repartido en dos hojas. Se fusionan
por nombre: un vendor que cubre Kansas City cubre los dos lados de la línea.
"""
import openpyxl, re, json, sys, collections

sys.stdout.reconfigure(encoding="utf-8")

XLSX = "Updated City_County Unit List.xlsx"
OUT = "markets.js"

US_STATES = {
    "Alabama":"AL","Alaska":"AK","Arizona":"AZ","Arkansas":"AR","California":"CA",
    "Colorado":"CO","Connecticut":"CT","Delaware":"DE","Florida":"FL","Georgia":"GA",
    "Hawaii":"HI","Idaho":"ID","Illinois":"IL","Indiana":"IN","Iowa":"IA","Kansas":"KS",
    "Kentucky":"KY","Louisiana":"LA","Maine":"ME","Maryland":"MD","Massachusetts":"MA",
    "Michigan":"MI","Minnesota":"MN","Mississippi":"MS","Missouri":"MO","Montana":"MT",
    "Nebraska":"NE","Nevada":"NV","New Hampshire":"NH","New Jersey":"NJ",
    "New Mexico":"NM","New York":"NY","North Carolina":"NC","North Dakota":"ND",
    "Ohio":"OH","Oklahoma":"OK","Oregon":"OR","Pennsylvania":"PA","Rhode Island":"RI",
    "South Carolina":"SC","South Dakota":"SD","Tennessee":"TN","Texas":"TX","Utah":"UT",
    "Vermont":"VT","Virginia":"VA","Washington":"WA","West Virginia":"WV",
    "Wisconsin":"WI","Wyoming":"WY","Washington DC":"DC","District of Columbia":"DC",
}
STATE_NAMES = sorted(US_STATES, key=len, reverse=True)

ZIP_RE = re.compile(r"^(\d{5})(-\d{4})?$")

def norm(v):
    if v is None: return ""
    s = str(v).strip()
    return s[:-2] if s.endswith(".0") else s

def is_zip(t):   return bool(ZIP_RE.match(t))
def zip5(t):     return ZIP_RE.match(t).group(1)
def is_count(t): return bool(re.fullmatch(r"[\d.,]+", t)) and not is_zip(t)
def is_county(t):
    return t.endswith(" County") or t.endswith(" Parish") or t.endswith(" city")

# Títulos descriptivos del informe que no sirven como nombre de mercado.
RENAME = {
    "Properties in TN market but not the state": "Tennessee (out of state)",
    "Properties in the KC Market": "Kansas City",
    "Properties in Chicago but in the Columbia Market": "Columbia - St Louis",
}

def clean_market(n):
    n = RENAME.get(n.strip(), n)
    n = re.sub(r"^HomeRiver Group\s*-\s*", "", n).strip()
    n = re.sub(r"\s*Market$", "", n).strip()
    return n or "Sin nombre"

def slug(s):
    return re.sub(r"[^A-Za-z0-9]+", "-", s).strip("-").lower()[:44]

def state_label(text):
    """Sigla SOLO si la fila es una etiqueta de estado de verdad.

    Regla estricta a propósito. Buscar el nombre de un estado en cualquier
    parte del texto falla de forma silenciosa y grave:
      · 'HomeRiver Group - Kansas City' contiene 'Kansas' -> etiquetaba los
        condados de Missouri como KS
      · 'Tennessee properties in Georgia' contiene 'Tennessee' -> etiquetaba
        condados de Georgia como TN
    Solo aceptamos la fila si ES el estado ('Arkansas') o si es exactamente
    '<Estado> Properties' ('Kentucky Properties'), que es como el informe
    introduce una sección de otro estado.
    """
    t = text.strip()
    if t in US_STATES:
        return US_STATES[t]
    m = re.fullmatch(r"(.+?)\s+Properties", t, flags=re.IGNORECASE)
    if m and m.group(1) in US_STATES:
        return US_STATES[m.group(1)]
    return None

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

markets = collections.OrderedDict()   # nombre limpio -> datos
caption_states = []
anomalies = []                   # para informar al usuario

for sheet in wb.sheetnames:
    ws = wb[sheet]
    rows = [[norm(c) for c in r] for r in ws.iter_rows(values_only=True)]
    if not rows: continue
    ncols = max((len(r) for r in rows), default=0)
    used = [any(len(r) > c and r[c] for r in rows) for c in range(ncols)]

    blocks, c = [], 0
    while c < ncols:
        if used[c]: blocks.append(c); c += 2
        else: c += 1

    sheet_state = US_STATES.get(sheet)

    for b in blocks:
        labels = [r[b] for r in rows if len(r) > b and r[b] and not is_count(r[b])]
        if not labels: continue

        # ── Clasificación en DOS PASADAS ──
        # Una sola pasada no basta porque hay nombres ambiguos:
        #   'Washington'  es un estado, pero también una ciudad de PA y de DC
        #   'New Market'  parece un mercado, pero es una ciudad de Alabama
        # Lo que desambigua es SIEMPRE lo que viene después:
        #   estado  -> le sigue un condado      ciudad  -> le siguen ZIPs
        n = len(labels)
        types = [None] * n

        # Pasada 1: lo que se reconoce por su propia forma.
        for i, t in enumerate(labels):
            if is_zip(t):      types[i] = "zip"
            elif is_county(t): types[i] = "county"

        def next_type(i):
            return types[i + 1] if i + 1 < n else None

        # Pasada 2: etiquetas de estado. Dos formas válidas:
        #   a) '<Estado> Properties'  -> inequívoco, ninguna ciudad se llama así
        #   b) el nombre exacto de un estado Y le sigue un condado
        #      (el 'Y le sigue un condado' es lo que evita confundir la ciudad
        #       de Washington, en PA y en DC, con el estado de Washington)
        for i, t in enumerate(labels):
            if types[i]: continue
            s = state_label(t)
            if not s: continue
            if re.fullmatch(r".+\s+Properties", t, flags=re.IGNORECASE):
                types[i] = "state"
            elif next_type(i) == "county":
                types[i] = "state"

        # Pasada 3: mercado (le sigue condado o estado) / ciudad (le siguen ZIPs).
        for i, t in enumerate(labels):
            if types[i]: continue
            nt = next_type(i)
            if nt in ("county", "state"): types[i] = "market"
            elif nt == "zip":             types[i] = "city"
            else:                         types[i] = "caption"   # título, se ignora

        # ── Recorrido ──
        block_state = sheet_state
        mkey = county = city = last_caption = None

        for i, t in enumerate(labels):
            k = types[i]

            if k == "caption":
                last_caption = t

            elif k == "state":
                block_state = state_label(t)
                if block_state != sheet_state:
                    caption_states.append(f"{sheet}: sección '{t}' -> estado {block_state}")
                # Secciones así (Arkansas/Mississippi en la hoja de Tennessee)
                # no traen fila de mercado: se usa el título del bloque.
                if not mkey:
                    name = clean_market(last_caption or t)
                    mkey = name
                    markets.setdefault(name, {"name": name,
                                              "counties": collections.OrderedDict()})
                county = city = None

            elif k == "market":
                name = clean_market(t)
                mkey = name
                markets.setdefault(name, {"name": name,
                                          "counties": collections.OrderedDict()})
                county = city = None

            elif k == "county":
                county = t; city = None
                if mkey:
                    markets[mkey]["counties"].setdefault(
                        county, {"state": block_state,
                                 "cities": collections.OrderedDict()})

            elif k == "city":
                city = t
                if mkey and county:
                    markets[mkey]["counties"][county]["cities"].setdefault(city, set())

            elif k == "zip":
                if mkey and county and city:
                    markets[mkey]["counties"][county]["cities"][city].add(zip5(t))
                else:
                    anomalies.append(
                        f"{sheet}: ZIP {t} huérfano "
                        f"(mercado={mkey}, condado={county}, ciudad={city})")

# ── Construir la salida ordenada ──
out = collections.OrderedDict()
all_zip = set()

for name in sorted(markets):
    m = markets[name]
    counties = collections.OrderedDict()
    states = set()
    for cty in sorted(m["counties"]):
        info = m["counties"][cty]
        cities = collections.OrderedDict()
        for city in sorted(info["cities"]):
            z = sorted(info["cities"][city])
            if z:
                cities[city] = z
                all_zip.update(z)
        if cities:
            counties[cty] = {"state": info["state"], "cities": cities}
            states.add(info["state"])
    if counties:
        out[slug(name)] = {"name": name,
                           "states": sorted(s for s in states if s),
                           "counties": counties}

nz = sum(len(z) for m in out.values() for c in m["counties"].values()
         for z in c["cities"].values())

print("=== RESULTADO FINAL ===")
print(f"  mercados   : {len(out)}")
print(f"  condados   : {sum(len(m['counties']) for m in out.values())}")
print(f"  ciudades   : {sum(len(c['cities']) for m in out.values() for c in m['counties'].values())}")
print(f"  ZIP únicos : {len(all_zip)}")
print()
for mid, m in out.items():
    n = sum(len(z) for c in m["counties"].values() for z in c["cities"].values())
    print(f"  {'/'.join(m['states']):<8} {m['name'][:30]:<30} {len(m['counties']):>3} cond {n:>5} ZIP")

if caption_states:
    print("\n=== ESTADO DEDUCIDO DE UN TÍTULO DE SECCIÓN (revísalo) ===")
    for c in sorted(set(caption_states)): print("  ·", c)

# ── Escribir el .js ──
def js(o): return json.dumps(o, ensure_ascii=False, separators=(",", ":"))

lines = ["""/* =========================================================================
   markets.js — MERCADOS, CONDADOS, CIUDADES y CÓDIGOS POSTALES
   -------------------------------------------------------------------------
   Generado desde 'Updated City_County Unit List.xlsx' (lista de unidades de
   HomeRiver Group). Sustituye al antiguo data/locations.js, que traía los 50
   estados con todos sus condados y ciudades: el sitio ofrecía cobertura en
   sitios donde la empresa no tiene ni una propiedad.

   La unidad de cobertura de un vendor es el CÓDIGO POSTAL de 5 dígitos.
   La jerarquía mercado -> condado -> ciudad existe solo para poder navegar
   y seleccionar en bloque; lo que se guarda en el perfil son los ZIP.

   Los ZIP+4 del archivo original (14.529) se recortaron a 5 dígitos: un
   ZIP+4 identifica una manzana concreta y nadie puede elegir entre 14.529
   casillas. Quedan 2.732 códigos.

   Mercados que cruzan la frontera estatal (Kansas City, Chattanooga,
   Columbia-St Louis, Ohio) están fusionados en uno solo, porque así los
   opera el negocio: quien cubre Kansas City cubre los dos lados.

   Estructura:
     MARKETS[id] = {
       name, states: ['KS','MO'],
       counties: { 'Johnson County': { state:'KS', cities: { 'Olathe': ['66061', …] } } }
     }
   ========================================================================= */
window.MARKETS = {"""]

items = []
for mid, m in out.items():
    cs = []
    for cty, info in m["counties"].items():
        cities = ",".join(f"{js(city)}:{js(z)}" for city, z in info["cities"].items())
        cs.append(f'{js(cty)}:{{"state":{js(info["state"])},"cities":{{{cities}}}}}')
    items.append(f'{js(mid)}:{{"name":{js(m["name"])},"states":{js(m["states"])},'
                 f'"counties":{{{",".join(cs)}}}}}')
lines.append(",\n".join(items))
lines.append("};\n")

with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))
print(f"\n{OUT} escrito")
